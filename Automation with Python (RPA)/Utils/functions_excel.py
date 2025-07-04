import os
import time

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

from Utils.integrated_logger import IntegratedLogger


def open_excel_file_to_dataframe(input_file_path, logger):
    """ 
    Abre o arquivo excel em um DataFrame, faz modificações necessárias para o projeto e retona o DataFrame 
    
    Parâmetros:
        input_file_path (str): Caminho do arquivo Excel a ser aberto.
    
    Retorna:
        pd.DataFrame: DataFrame com os dados do arquivo Excel e as modificações realizadas.
    
    Raises:
        FileNotFoundError: Se o arquivo não for encontrado no caminho especificado.
        Exception: Para qualquer outro erro que ocorra durante o processo.
    """
    try:
        logger.info("Iniciando abertura do arquivo Excel e retornando um DataFrame")

        # Verifica se o arquivo existe
        if not os.path.exists(input_file_path):
            logger.debug(" arquivo não foi encontrado.")
            raise FileNotFoundError(f"O arquivo {input_file_path} não foi encontrado.") 
        logger.info(f"O arquivo de Excel com os dados de entrada foi encontrado.")
        logger.debug(f"O arquivo foi encontrado na pasta indicada: {input_file_path}")
        
        # Abre o arquivo excel em DataFrame, na pasta 'Groupo 1' e preenche valores vazios como 'NA'
        df_input = pd.read_excel(input_file_path, "Grupo 1 ", na_values=["NA"],dtype=object)
        logger.info("DataFrame com base no arquivo Excel criado com sucesso")

        return df_input
    
    except Exception as erro:
        logger.error('Execução open_excel_file_to_dataframe')
        # Para o processo para depuração manual
        raise


def create_output_dataframe(df_input, logger):
    """
    Cria um DataFrame vazio com colunas predefinidas.
    
    Retorna:
        pd.DataFrame: DataFrame com as colunas predefinidas.
    
    Raises:
        Exception: Para qualquer erro que ocorra durante o processo.
    """
    try:
        logger.info("Iniciando a criação do DataFrame para receber os dados de saída")
        
        # Definindo colunas
        columns = [
            "CNPJ", "RAZÃO SOCIAL", "NOME FANTASIA",
            "ENDEREÇO", "CEP", "DESCRIÇÃO MATRIZ FILIAL",
            "TELEFONE + DDD", "E-MAIL", "VALOR DO PEDIDO",
            "DIMENSÕES CAIXA (altura x largura x comprimento cm)", "PESO DO PRODUTO", "TIPO DE SERVIÇO JADLOG",
            "TIPO DE SERVIÇO CORREIOS", "VALOR COTAÇÃO JADLOG", "VALOR COTAÇÃO CORREIOS",
            "PRAZO DE ENTREGA CORREIOS", "STATUS"
        ]
        # Criando o DataFrame com as colunas definidas
        df_output = pd.DataFrame(columns=columns,dtype=str)
        logger.info("DataFrame para receber as saídas criado com sucesso.")

        # Alimenta o DataFrame de saída com a informações já existentes
        df_output['CNPJ'] = df_input['CNPJ']
        df_output = df_output.set_index('CNPJ')
        df_input = df_input.set_index('CNPJ')
        df_output.update(df_input)
        df_output = df_output.reset_index()
        
        return df_output
    
    except Exception as erro:
        logger.error("Execução de create_output_dataframe")
        # Para o processo para depuração manual
        raise


def save_df_output_to_excel(output_path, df_output, logger):
    """
    Salva o DataFrame em um arquivo Excel no caminho especificado.
    
    Parâmetros:
        output_path (str): Caminho onde o arquivo Excel será salvo.
        cnpj (str): CNPJ que será usado no nome do arquivo.
        df_output (pd.DataFrame): DataFrame a ser salvo.
    
    Retorna:
        str: Caminho do arquivo Excel gerado.
    
    Raises:
        Exception: Para qualquer erro que ocorra durante o processo.
    """
    try:
        logger.info("Iniciando a criação da planilha Excel com os dados de saída")

        # Gerando nome do arquivo baseado na data e hora atual
        current_date = time.strftime("%Y-%m-%d_%Hh%Mm%Ss")
        file_name = f"cnpj_{current_date}.xlsx"
        logger.debug(f"Nome do arquivo criado: {file_name}")

        # Salvando DataFrame como arquivo Excel
        output_file_path = f"{output_path}/{file_name}"
        df_output.to_excel(output_file_path, index=False)
        logger.info(f"Sucesso, arquivo criado: {file_name}")
        logger.debug(f"Arquivo Excel criado com sucesso em: {output_path}")

        return output_file_path

    except Exception as erro:
        logger.error('Execução save_df_output_excel')
        # Para o processo para depuração manual
        raise


def clean_df_if_null(df_to_clean, na_not_allowed_columns, logger):
    """
    Remove linhas com células vazias de um DataFrame e registra os CNPJs e as colunas com células vazias.
    
    Args:
        df_to_clean (pd.DataFrame): DataFrame a ser limpo.
        na_not_allowed_columns (list): Lista com as colunas que não podem estar em branco
    
    Returns:
        pd.DataFrame: DataFrame limpo.
        list: Lista de dicionários com CNPJs e colunas com células vazias.
    
    Raises:
        Exception: Para qualquer erro que ocorra durante o processo.
    """
    try:
        logger.info("Iniciando o processo de limpeza das células vazias do DataFrame")
        # Cria uma lista para receber os CNPJs com alguma célula vazia
        empty_cells = []

        # Verifica se há células em branco em cada linha
        rows_to_drop = []        
        for _, row in df_to_clean.iterrows():
            empty_rows = row[row.isnull()].index.tolist()
            # Registra as linhas vazias na lista como um dicionário
            if empty_rows:
                empty_cells.append({
                    "CNPJ": row["CNPJ"],
                    "NA": empty_rows
                    })
            # Verifica se há colunas com células vazias que estão na lista de colunas onde valores nulos não são permitidos
                if any(
                    column in na_not_allowed_columns
                    # Para cada coluna com célula vazia na linha atual, 
                    # verifica se a coluna está na lista de colunas que não permitem valores nulos
                    for column in empty_rows
                    ):
                    rows_to_drop.append(row.name)

        # Registra os CNPJs com células vazias
        if empty_cells:
            logger.info(f"CNPJs com células vazias: {empty_cells}")
        else:
            logger.info("Nenhuma célula vazia encontrada.")

        # Remove as linhas que contêm células em branco
        df_clean = df_to_clean.dropna()
        logger.info("Células vazias/NA removidas do DataFrame")
        
        return df_clean, empty_cells

    except Exception as erro:
        logger.error('Execução clean_df_if_null')
        # Para o processo para depuração manual
        raise


def write_if_null_output(df_output, empty_cells, logger):
    """
    Atualiza a coluna 'Status' no DataFrame de saída com informações sobre células vazias.

    Parâmetros:
    df_output (pd.DataFrame): DataFrame que será atualizado.
    empty_cells (list): Lista de dicionários contendo CNPJs e colunas vazias.

    Retorna:
    pd.DataFrame: DataFrame atualizado com a coluna 'Status' preenchida.
    """
    try:
        logger.info("Iniciando o processo de registro de células vazias")

        # Itera sobre a lista de células vazias para preencher a coluna 'Status'
        for empty in empty_cells:
            # Verifica se, na linha, o índice CNPJ é igual
            value = empty['CNPJ']
            index = df_output[df_output["CNPJ"] == value].index
            # Verifica se a linha com o CNPJ existe no df_output e preenche
            if not index.empty:
                # Atualiza a coluna 'Status' na linha correspondente
                df_output.at[index, "STATUS"] = f"Campos vazios: {empty['NA']}"
                logger.info(f"Coluna 'Status' atualizada para o CNPJ {empty['CNPJ']}: {empty['NA']}")

        return df_output


    except Exception as erro:
        logger.error('Execução write_if_null_output')
        # Para o processo para depuração manual
        raise    


def compare_quotation(df_output: pd.DataFrame, output_file_path: str, logger) -> None:
    """
    Compara os valores de cotação entre Correios e Jadlog e destaca no Excel o menor valor com uma cor visual.

    A função cria uma coluna auxiliar para identificar o menor valor entre as cotações de cada linha
    e aplica preenchimento verde à célula correspondente no arquivo Excel. Em seguida, remove a
    coluna temporária e salva o arquivo.

    Parâmetros:
        df_output (pd.DataFrame): DataFrame contendo as colunas de cotação dos Correios e da Jadlog.
        output_file_path (str): Caminho absoluto do arquivo Excel a ser editado e salvo.
        logger: Instância de logger para registrar as etapas e falhas da execução.

    Retorna:
        None

    Raises:
        ValueError: Se as colunas esperadas de cotação não existirem no DataFrame.
        Exception: Para qualquer outro erro ocorrido durante o processo.
    """
    try:
        logger.info("Iniciando o processo de comparação de cotações: Correios vs Jadlog")

        # Verifica se as colunas esperadas existem no DataFrame
        colunas_esperadas = ["VALOR COTAÇÃO CORREIOS", "VALOR COTAÇÃO JADLOG"]
        for coluna in colunas_esperadas:
            if coluna not in df_output.columns:
                logger.error(f"Coluna ausente: {coluna}")
                raise ValueError(f"A coluna '{coluna}' não está presente no DataFrame.")

        # Cria cópia para transformar os valores em float para comparação
        df_formatado = df_output.copy()
        for col in colunas_esperadas:
            df_formatado[col] = (
                df_formatado[col]
                .astype(str)
                .str.replace("R$", "", regex=False)
                .str.replace(",", ".", regex=False)
                .astype(float)
            )

        # Identifica o nome da coluna com o menor valor para cada linha
        coluna_tmp = "__MENOR_VALOR_TMP__"
        df_output[coluna_tmp] = df_formatado[colunas_esperadas].idxmin(axis=1)
        logger.info(f"Coluna auxiliar '{coluna_tmp}' criada para identificar o menor valor por linha.")

        # Carrega o arquivo Excel existente
        workbook = load_workbook(output_file_path)
        worksheet = workbook.active
        logger.info("Arquivo Excel carregado com sucesso.")

        # Define cor verde para preenchimento da célula
        preenchimento_verde = PatternFill(start_color="33CC33", end_color="33CC33", fill_type="solid")

        for index, row in df_output.iterrows():
            coluna_menor = row[coluna_tmp]
            coluna_excel = df_output.columns.get_loc(coluna_menor) + 1  # +1 porque Excel é 1-indexed
            linha_excel = index + 2  # +2 porque a primeira linha é o cabeçalho

            celula = worksheet.cell(row=linha_excel, column=coluna_excel)
            if celula.value is not None:
                celula.fill = preenchimento_verde
                logger.debug(f"Célula com menor valor preenchida na linha {linha_excel}, coluna {coluna_excel}.")

        # Remove a coluna temporária usada apenas internamente
        df_output.drop(columns=coluna_tmp, inplace=True)
        logger.debug(f"Coluna auxiliar '{coluna_tmp}' removida do DataFrame.")

        # Salva o Excel com as alterações
        workbook.save(output_file_path)
        logger.info(f"Arquivo Excel atualizado com destaques salvos em: {output_file_path}")

    except Exception as erro:
        logger.error("Erro na execução da função 'compare_quotation'", exc_info=True)
        raise


def make_column_endereco(df, logger):
    """
    Combina as colunas de logradouro, número e município em uma única coluna de endereço.
    
    Args:
        df (pd.DataFrame): DataFrame original contendo as colunas 'LOGRADOURO', 'NÚMERO' e 'MUNICÍPIO'.
    
    Returns:
        pd.DataFrame: DataFrame com a nova coluna 'ENDEREÇO' e sem as colunas originais.
    
    """
    
    endereco_cols = ["LOGRADOURO", "NÚMERO", "MUNICÍPIO"]
    
    try:
        # Verifica se todas as colunas necessárias estão presentes no DataFrame
        if not all(col in df.columns for col in endereco_cols):
            raise ValueError("O DataFrame deve conter as colunas 'LOGRADOURO', 'NÚMERO' e 'MUNICÍPIO'.")
        
        # Cria a nova coluna 'ENDEREÇO' concatenando as colunas necessárias
        df["ENDEREÇO"] = df[endereco_cols].agg(lambda x: ', '.join(x.dropna().astype(str)), axis=1)
        
        # Remove as colunas originais que foram combinadas
        df.drop(columns=endereco_cols, inplace=True)
        return df
    
    except Exception as erro:
        logger.error('Execução make_endereco')
        # Para o processo para depuração manual
        raise   


def make_jadlog_correios_dataframes(df_output, api_data, logger):
    """
    Pega as informações do Dataframe original, junta com as informações dadas pela API 
    e divide em três dataframes que serão usados no fluxo principal.
    
    Args:
        df_output (pd.DataFrame): Dataframe onde ocorrem as edições.
        api_data (pd.DataFrame): Dataframe contendo as informações retornadas pela API.
        logger (IntegratedLogger): Logger usado para gerar arquivos de log.
    
    Returns:
        tuple: Três dataframes - o de saída editado, o utilizado no processo dos correios e o utilizado no jadlog.

    """
    
    correios_columns = ['CNPJ','DIMENSÕES CAIXA (altura x largura x comprimento cm)',
                        'PESO DO PRODUTO','TIPO DE SERVIÇO CORREIOS',
                        'CEP']
    jadlog_columns = ['CNPJ','TIPO DE SERVIÇO JADLOG',
                    'DIMENSÕES CAIXA (altura x largura x comprimento cm)','PESO DO PRODUTO',
                    'CEP','VALOR DO PEDIDO']
    
    try:
        # Ajusta os índices dos DataFrames para CNPJ
        df_output = df_output.set_index('CNPJ')
        api_data = api_data.set_index('CNPJ')
        
        # Atualiza df_output com dados da API
        df_output.update(api_data)
        df_output = df_output.reset_index()
        
        # Limpa e prepara DataFrame para Correios
        df_correios, empty_cells = clean_df_if_null(df_output[correios_columns], correios_columns, logger)
        df_output = write_if_null_output(df_output, empty_cells, logger)
        
        # Limpa e prepara DataFrame para JadLog
        df_jadlog, empty_cells = clean_df_if_null(df_output[jadlog_columns], jadlog_columns, logger)
        df_jadlog = df_jadlog.astype(str)
        df_jadlog['VALOR DO PEDIDO'] = df_jadlog['VALOR DO PEDIDO'].apply(lambda x: x.replace('.', ','))
        
        # Registra as células vazias no DataFrame de saída
        df_output = write_if_null_output(df_output, empty_cells, logger)
        
        return df_output, df_correios, df_jadlog

    except Exception as erro:
        logger.error('Execução make_jadlog_correios_dataframes')
        # Para o processo para depuração manual
        raise   